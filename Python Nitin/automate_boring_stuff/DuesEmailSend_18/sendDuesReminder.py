# sendDuesReminder.py - Sends emails based on the payment status in spreadsheet
import openpyxl, smtplib, sys, pyinputplus

gmailSmtpURL = "smtp.gmail.com"
senderUsername = str(input("Enter your username: "))
try: senderPasswd = pyinputplus.inputPassword("Enter your password: ", limit = 3)
except pyinputplus.RetryLimitException: print("Retry limit reached !"); exit()
wb = openpyxl.load_workbook("dues.xlsx")
sheet1 = wb["Sheet1"]
# print(sheet1)
lastCol = sheet1.max_column
latestMonth = sheet1.cell(row = 1, column = lastCol).value
unpaidMembers = {}
for r in range(2, sheet1.max_row):
    payment = sheet1.cell(row = r, column = lastCol).value
    if payment != "paid":
        name = sheet1.cell(row = r, column = 1).value
        email = sheet1.cell(row = r, column = 2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP(gmailSmtpURL, 587)
smtpObj.ehlo()
smtpObj.starttls()
print(senderUsername + "@gmail.com", senderPasswd)
smtpObj.login(senderUsername + "@gmail.com", senderPasswd)

for name, email in unpaidMembers.values():
    body = "Subject: %s dues unpaid.\nDear %s,\nTesting email sent to you VERBOSE: %s. Thank you!'" % (latestMonth, name, latestMonth)
    print("sending email to %s...", name)
    status = smtpObj.sendmail(senderUsername + "@gmail.com", email, body)
    if status != {}:
        print("There's problem in sending email to: %s (%s)" % (email, name))
smtpObj.quit()